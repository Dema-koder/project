from io import BytesIO
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def export_shopping_list_pdf(event, shopping_list, items):
    """Экспорт списка покупок в PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    # Стили
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=16,
        spaceAfter=30
    )

    # Заголовок
    elements.append(Paragraph(f"Список покупок: {event.name}", title_style))
    elements.append(Paragraph(f"Дата: {event.event_date}", styles['Normal']))
    elements.append(Paragraph(f"Гостей: {event.number_of_guests}", styles['Normal']))
    elements.append(Spacer(1, 20))

    # Группируем по категориям
    categories = {}
    for item in items:
        category = item.ingredient.category or 'Другое'
        if category not in categories:
            categories[category] = []
        categories[category].append(item)

    # Создаем таблицы для каждой категории
    for category, category_items in categories.items():
        elements.append(Paragraph(category, styles['Heading2']))

        table_data = [['Продукт', 'Количество', 'Примерная стоимость', 'Куплено']]

        for item in category_items:
            table_data.append([
                item.ingredient.name,
                f"{item.quantity_needed} {item.ingredient.get_unit_display()}",
                f"{item.estimated_cost:.2f} ₽" if item.estimated_cost else "-",
                "✓" if item.purchased else "☐"
            ])

        # Создаем таблицу
        table = Table(table_data, colWidths=[3*inch, 1.5*inch, 1.5*inch, 0.8*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 20))

    # Итоговая стоимость
    elements.append(Paragraph(
        f"Общая ориентировочная стоимость: {shopping_list.total_cost:.2f} ₽",
        styles['Heading2']
    ))

    # Создаем PDF
    doc.build(elements)
    buffer.seek(0)

    from django.http import HttpResponse
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="shopping_list_{event.name}.pdf"'
    return response

def export_shopping_list_excel(event, shopping_list, items):
    """Экспорт списка покупок в Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Список покупок"

    # Заголовок
    ws['A1'] = f"Список покупок: {event.name}"
    ws['A2'] = f"Дата: {event.event_date}"
    ws['A3'] = f"Гостей: {event.number_of_guests}"

    # Стили заголовка
    title_font = Font(size=14, bold=True)
    ws['A1'].font = title_font

    # Заголовки таблицы
    headers = ['Категория', 'Продукт', 'Количество', 'Единица', 'Примерная стоимость', 'Куплено']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Данные
    row = 6
    for item in items:
        ws.cell(row=row, column=1, value=item.ingredient.category or 'Другое')
        ws.cell(row=row, column=2, value=item.ingredient.name)
        ws.cell(row=row, column=3, value=item.quantity_needed)
        ws.cell(row=row, column=4, value=item.ingredient.get_unit_display())
        ws.cell(row=row, column=5, value=float(item.estimated_cost) if item.estimated_cost else None)
        ws.cell(row=row, column=6, value='Да' if item.purchased else 'Нет')
        row += 1

    # Итог
    ws.cell(row=row+1, column=4, value="Общая стоимость:")
    ws.cell(row=row+1, column=5, value=float(shopping_list.total_cost))
    ws.cell(row=row+1, column=5).number_format = '#,##0.00 ₽'

    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Сохраняем в буфер
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    from django.http import HttpResponse
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="shopping_list_{event.name}.xlsx"'
    return response